"""Generic multi-university crawler and programme extractor.

Reads target_universities.csv and writes output_programmes.xlsx with candidate
pages, extracted programme review rows, and a run log. Extraction is heuristic
and intentionally generic so it can work across different university websites.
"""
from __future__ import annotations

import argparse
import csv
import os
import re
import time
from collections import deque
from dataclasses import dataclass
from datetime import datetime, timezone
from difflib import SequenceMatcher
from typing import Callable, Dict, List, Optional, Set, Tuple
from urllib.parse import urldefrag, urljoin, urlparse
from urllib.robotparser import RobotFileParser

import pandas as pd
import requests
from openpyxl.styles import PatternFill
from openpyxl.utils import get_column_letter
from bs4 import BeautifulSoup

INPUT_FILE = "target_universities.csv"
OUTPUT_FILE = "output_programmes.xlsx"
CRAWL_DEPTH_LIMIT = 2
MAX_PAGES_PER_UNIVERSITY = 50
REQUEST_DELAY_SECONDS = 1.0
MIN_CANDIDATE_SCORE = 20
USER_AGENT = "GenericProgrammeCandidateCrawler/1.0 (+https://example.local; polite educational crawler)"

INPUT_COLUMNS = ["University Name", "Country", "Seed URL", "Allowed Domain", "Notes"]
CANDIDATE_COLUMNS = [
    "University Name", "Country", "URL", "Page Title", "Candidate Score",
    "Matched Keywords", "Candidate Type", "Reason", "Needs Review",
]
PROGRAMME_COLUMNS = [
    "University Name", "Country", "Programme Name", "Programme Type",
    "Target Students", "Language", "Duration / Period", "Programme Dates",
    "Application Deadline", "Programme Fee", "Housing", "Credits / Certificate",
    "Main Contents", "Eligibility", "Source URL", "Source Page Title",
    "Confidence Score", "Missing Fields", "Review Status", "Duplicate Group",
    "Duplicate Status", "Last Checked", "Extraction Method", "Notes",
]
RUN_LOG_COLUMNS = [
    "Start Time", "End Time", "Universities Processed",
    "Universities With Candidates", "Universities With Extracted Programmes",
    "Total Pages Visited", "Total Candidate Pages", "Candidate Pages Read",
    "Total Programme Rows", "Fallback Rows", "Duplicate Rows", "Status",
    "Warnings", "Errors",
]

HIGH_PRIORITY_KEYWORDS = [
    "short-term program", "short-term programme", "summer program", "summer programme",
    "inbound program", "inbound programme", "japanese language program",
    "japanese language programme", "customized program", "customised programme",
    "study abroad", "exchange program", "non-degree", "certificate",
    "application deadline", "program fee", "programme fee", "tuition", "housing",
    "accommodation", "eligibility",
]
LOW_PRIORITY_KEYWORDS = ["international students", "admissions", "campus life", "news", "events"]
URL_HINTS = [
    "short", "summer", "inbound", "japanese", "language", "custom", "exchange",
    "non-degree", "certificate", "program", "programme", "study-abroad", "international",
]
AVOID_URL_PARTS = [
    "/login", "signin", "auth", "wp-login", "logout", "contact", "inquiry",
    "calendar", "tag/", "category/", "author/", "share=", "?replytocom=",
]
NEGATIVE_URL_EXTENSIONS = (
    ".jpg", ".jpeg", ".png", ".gif", ".svg", ".webp", ".zip", ".doc", ".docx",
    ".ppt", ".pptx", ".xls", ".xlsx", ".css", ".js", ".mp4", ".mp3",
)

IMPORTANT_REVIEW_FIELDS = [
    "Programme Name", "Programme Dates", "Application Deadline", "Programme Fee",
    "Housing", "Eligibility", "Source URL",
]


def is_missing_value(value: object) -> bool:
    return not str(value or "").strip() or str(value).strip().lower() in {"unknown", "n/a", "none"}


def normalise_duplicate_text(value: str) -> str:
    value = re.sub(r"https?://", "", value.lower())
    value = re.sub(r"[^a-z0-9]+", " ", value)
    words = [word for word in value.split() if word not in {"program", "programme", "the", "and", "of"}]
    return " ".join(words).strip()


@dataclass
class UniversityTarget:
    name: str
    country: str
    seed_url: str
    allowed_domain: str
    notes: str


@dataclass
class CandidatePage:
    university_name: str
    country: str
    url: str
    title: str
    score: int
    matched_keywords: List[str]
    candidate_type: str
    reason: str
    needs_review: str = "Yes"


@dataclass
class ProgrammeRow:
    university_name: str
    country: str
    programme_name: str
    programme_type: str
    target_students: str
    language: str
    duration_period: str
    programme_dates: str
    application_deadline: str
    programme_fee: str
    housing: str
    credits_certificate: str
    main_contents: str
    eligibility: str
    source_url: str
    source_page_title: str
    confidence_score: int
    missing_fields: str
    review_status: str
    duplicate_group: str
    duplicate_status: str
    last_checked: str
    extraction_method: str
    notes: str


def normalise_url(url: str) -> str:
    clean, _ = urldefrag(str(url).strip())
    parsed = urlparse(clean)
    if not parsed.scheme:
        clean = "https://" + clean
        parsed = urlparse(clean)
    return parsed._replace(fragment="").geturl().rstrip("/")


def normalise_domain(domain_or_url: str) -> str:
    value = str(domain_or_url).strip().lower()
    parsed = urlparse(value if "://" in value else f"https://{value}")
    return (parsed.hostname or value).removeprefix("www.")


def is_allowed_url(url: str, allowed_domain: str) -> bool:
    parsed = urlparse(url)
    host = (parsed.hostname or "").lower().removeprefix("www.")
    domain = normalise_domain(allowed_domain)
    return parsed.scheme in {"http", "https"} and (host == domain or host.endswith(f".{domain}"))


def should_skip_url(url: str) -> bool:
    lower = url.lower()
    return (
        lower.startswith(("mailto:", "tel:", "javascript:"))
        or lower.endswith(NEGATIVE_URL_EXTENSIONS)
        or any(part in lower for part in AVOID_URL_PARTS)
    )


def get_robot_parser(seed_url: str) -> RobotFileParser:
    parsed = urlparse(seed_url)
    rp = RobotFileParser()
    rp.set_url(f"{parsed.scheme}://{parsed.netloc}/robots.txt")
    try:
        response = requests.get(rp.url, headers={"User-Agent": USER_AGENT}, timeout=10)
        if response.status_code < 500:
            response.raise_for_status()
            rp.parse(response.text.splitlines())
            return rp
    except requests.RequestException:
        pass
    rp.parse(["User-agent: *", "Allow: /"])
    return rp


def fetch_html(url: str, robot_parser: RobotFileParser) -> Optional[str]:
    if not robot_parser.can_fetch(USER_AGENT, url):
        return None
    response = requests.get(url, headers={"User-Agent": USER_AGENT}, timeout=20)
    response.raise_for_status()
    content_type = response.headers.get("Content-Type", "").lower()
    if "text/html" not in content_type and "application/xhtml" not in content_type:
        return None
    return response.text


def extract_page(html: str, page_url: str) -> Tuple[str, str, List[str]]:
    soup = BeautifulSoup(html, "html.parser")
    for tag in soup(["script", "style", "noscript", "svg", "form"]):
        tag.decompose()
    title = soup.title.get_text(" ", strip=True) if soup.title else ""
    text = re.sub(r"\s+", " ", soup.get_text(" ", strip=True))
    links = sorted({normalise_url(urljoin(page_url, anchor["href"])) for anchor in soup.find_all("a", href=True)})
    return title, text, links


def score_page(title: str, text: str, url: str) -> Tuple[int, List[str], str, str]:
    haystack = f"{title} {url} {text[:10000]}".lower()
    matched_high = [kw for kw in HIGH_PRIORITY_KEYWORDS if kw in haystack]
    matched_low = [kw for kw in LOW_PRIORITY_KEYWORDS if kw in haystack]
    url_matches = [hint for hint in URL_HINTS if hint in url.lower()]
    score = min(100, len(matched_high) * 15 + len(matched_low) * 5 + len(url_matches) * 3)
    if any(term in haystack for term in ["apply", "application", "deadline", "fee", "tuition", "eligibility"]):
        score = min(100, score + 10)
    matched = matched_high + matched_low
    if matched_high:
        candidate_type = "Strong programme candidate"
        reason = "High-priority programme keywords found."
    elif matched_low or url_matches:
        candidate_type = "Possible related page"
        reason = "Lower-priority keywords or URL hints found."
    else:
        candidate_type = "Low relevance"
        reason = "No configured candidate keywords found."
    return score, matched, candidate_type, reason


def should_enqueue_link(url: str, allowed_domain: str) -> bool:
    if should_skip_url(url) or not is_allowed_url(url, allowed_domain):
        return False
    lower = url.lower()
    return any(hint in lower for hint in URL_HINTS) or lower.count("/") <= 5


def crawl_university(target: UniversityTarget) -> Tuple[List[CandidatePage], int, List[str]]:
    seed_url = normalise_url(target.seed_url)
    allowed_domain = target.allowed_domain or normalise_domain(seed_url)
    robot_parser = get_robot_parser(seed_url)
    queue = deque([(seed_url, 0)])
    visited: Set[str] = set()
    candidates_by_url: Dict[str, CandidatePage] = {}
    warnings: List[str] = []

    while queue and len(visited) < MAX_PAGES_PER_UNIVERSITY:
        url, depth = queue.popleft()
        if url in visited or should_skip_url(url) or not is_allowed_url(url, allowed_domain):
            continue
        visited.add(url)
        try:
            html = fetch_html(url, robot_parser)
            time.sleep(REQUEST_DELAY_SECONDS)
            if html is None:
                continue
            title, text, links = extract_page(html, url)
            score, matched, candidate_type, reason = score_page(title, text, url)
            if score >= MIN_CANDIDATE_SCORE:
                candidates_by_url[url] = CandidatePage(
                    university_name=target.name,
                    country=target.country,
                    url=url,
                    title=title,
                    score=score,
                    matched_keywords=matched,
                    candidate_type=candidate_type,
                    reason=reason,
                )
            if depth < CRAWL_DEPTH_LIMIT:
                for link in links:
                    if link not in visited and should_enqueue_link(link, allowed_domain):
                        queue.append((link, depth + 1))
        except Exception as exc:  # keep processing other pages and universities
            warnings.append(f"{url}: {exc}")

    candidates = sorted(candidates_by_url.values(), key=lambda page: page.score, reverse=True)
    return candidates, len(visited), warnings



def extract_heading(html: str) -> str:
    soup = BeautifulSoup(html, "html.parser")
    for tag_name in ["h1", "h2"]:
        tag = soup.find(tag_name)
        if tag:
            heading = re.sub(r"\s+", " ", tag.get_text(" ", strip=True))
            if heading:
                return heading
    return ""


def infer_programme_name(title: str, heading: str, url: str) -> Tuple[str, bool]:
    candidates = [heading, title]
    for value in candidates:
        cleaned = re.sub(r"\s*[|–—-]\s*(.+University|Admissions|International.*)$", "", value, flags=re.I).strip()
        if cleaned and len(cleaned) >= 4:
            return cleaned[:200], value != heading or not heading
    path_tail = urlparse(url).path.rstrip("/").split("/")[-1]
    inferred = re.sub(r"[-_]+", " ", path_tail).strip().title()
    return (inferred or "Unknown Programme"), True


def first_match(text: str, patterns: List[str]) -> str:
    for pattern in patterns:
        match = re.search(pattern, text, flags=re.I)
        if match:
            return re.sub(r"\s+", " ", match.group(1)).strip(" :-–—")[:500]
    return "Unknown"


def detect_programme_type(text: str, url: str) -> str:
    haystack = f"{text} {url}".lower()
    mapping = [
        ("Summer Programme", ["summer program", "summer programme", "summer school"]),
        ("Short-term Programme", ["short-term", "short term"]),
        ("Exchange Programme", ["exchange program", "exchange programme"]),
        ("Japanese Language Programme", ["japanese language"]),
        ("Certificate Programme", ["certificate"]),
        ("Study Abroad Programme", ["study abroad"]),
    ]
    for label, terms in mapping:
        if any(term in haystack for term in terms):
            return label
    return "Unknown"


def extract_programme_from_candidate(target: UniversityTarget, page: CandidatePage, html: str) -> Optional[ProgrammeRow]:
    title, text, _ = extract_page(html, page.url)
    heading = extract_heading(html)
    name, inferred = infer_programme_name(title or page.title, heading, page.url)
    if not name or not page.url:
        return None

    language = "English" if re.search(r"\bEnglish\b", text, re.I) else ("Japanese" if re.search(r"\bJapanese\b", text, re.I) else "Unknown")
    fields = {
        "Programme Type": detect_programme_type(text, page.url),
        "Target Students": first_match(text, [r"Target(?:ed)? (?:Students|Participants|Audience)[:\s]+([^.;\n]{3,180})", r"Eligibility[:\s]+([^.;\n]{3,180})"]),
        "Language": language,
        "Duration / Period": first_match(text, [r"Duration[:\s]+([^.;\n]{3,180})", r"Period[:\s]+([^.;\n]{3,180})"]),
        "Programme Dates": first_match(text, [r"(?:Program|Programme) Dates?[:\s]+([^.;\n]{3,180})", r"Dates?[:\s]+([^.;\n]{3,180})"]),
        "Application Deadline": first_match(text, [r"Application Deadline[:\s]+([^.;\n]{3,180})", r"Deadline[:\s]+([^.;\n]{3,180})"]),
        "Programme Fee": first_match(text, [r"(?:Program|Programme) Fee[:\s]+([^.;\n]{3,180})", r"Tuition[:\s]+([^.;\n]{3,180})", r"Fee[:\s]+([^.;\n]{3,180})"]),
        "Housing": "Available" if re.search(r"\b(housing|accommodation|dormitory|residence)\b", text, re.I) else "Unknown",
        "Credits / Certificate": first_match(text, [r"(?:Credits?|Certificate)[:\s]+([^.;\n]{3,180})"]),
        "Main Contents": first_match(text, [r"(?:Overview|About|Contents?|Description)[:\s]+([^\n]{20,500})"]),
        "Eligibility": first_match(text, [r"Eligibility[:\s]+([^\n]{3,500})", r"Requirements?[:\s]+([^\n]{3,500})"]),
    }
    review_values = {
        "Programme Name": name,
        "Programme Dates": fields["Programme Dates"],
        "Application Deadline": fields["Application Deadline"],
        "Programme Fee": fields["Programme Fee"],
        "Housing": fields["Housing"],
        "Eligibility": fields["Eligibility"],
        "Source URL": page.url,
    }
    missing = [key for key in IMPORTANT_REVIEW_FIELDS if is_missing_value(review_values.get(key))]
    present_important = len(IMPORTANT_REVIEW_FIELDS) - len(missing)
    programme_keywords = sum(1 for kw in HIGH_PRIORITY_KEYWORDS if kw in text.lower())
    confidence = min(100, max(0, page.score + (12 if not inferred else -8) + 8 * present_important + min(10, programme_keywords * 2)))
    if page.score < MIN_CANDIDATE_SCORE or (present_important <= 2 and programme_keywords <= 1):
        review_status = "Low confidence"
    elif page.url and inferred and missing:
        review_status = "Needs human review"
    elif name and page.url and present_important >= 5:
        review_status = "Likely valid"
    else:
        review_status = "Needs human review"
    notes = "Programme name inferred from page title, heading, or URL." if inferred else ""
    return ProgrammeRow(target.name, target.country, name, fields["Programme Type"], fields["Target Students"], fields["Language"], fields["Duration / Period"], fields["Programme Dates"], fields["Application Deadline"], fields["Programme Fee"], fields["Housing"], fields["Credits / Certificate"], fields["Main Contents"], fields["Eligibility"], page.url, title or page.title, confidence, ", ".join(missing), review_status, "", "Unique", datetime.now(timezone.utc).strftime("%Y-%m-%d"), "heuristic", notes)


def make_fallback_programme_row(target: UniversityTarget, page: CandidatePage) -> ProgrammeRow:
    name, _ = infer_programme_name(page.title, "", page.url)
    missing = ["Programme Dates", "Application Deadline", "Programme Fee", "Housing", "Eligibility"]
    return ProgrammeRow(target.name, target.country, name, "Unknown", "Unknown", "Unknown", "Unknown", "Unknown", "Unknown", "Unknown", "Unknown", "Unknown", "Unknown", "Unknown", page.url, page.title, min(page.score, 45), ", ".join(missing), "Needs human review", "", "Unique", datetime.now(timezone.utc).strftime("%Y-%m-%d"), "fallback", "Fallback candidate row because no programme rows were extracted for this university. Programme name inferred from page title or URL.")


def programme_to_dict(row: ProgrammeRow) -> Dict[str, object]:
    return {
        "University Name": row.university_name, "Country": row.country, "Programme Name": row.programme_name,
        "Programme Type": row.programme_type, "Target Students": row.target_students, "Language": row.language,
        "Duration / Period": row.duration_period, "Programme Dates": row.programme_dates,
        "Application Deadline": row.application_deadline, "Programme Fee": row.programme_fee, "Housing": row.housing,
        "Credits / Certificate": row.credits_certificate, "Main Contents": row.main_contents, "Eligibility": row.eligibility,
        "Source URL": row.source_url, "Source Page Title": row.source_page_title, "Confidence Score": row.confidence_score,
        "Missing Fields": row.missing_fields, "Review Status": row.review_status, "Duplicate Group": row.duplicate_group,
        "Duplicate Status": row.duplicate_status, "Last Checked": row.last_checked, "Extraction Method": row.extraction_method,
        "Notes": row.notes,
    }


def mark_duplicate_programmes(rows: List[Dict[str, object]]) -> int:
    """Mark exact and near duplicate programme rows in-place."""
    groups: List[List[int]] = []
    signatures: List[Tuple[str, str]] = []
    for index, row in enumerate(rows):
        name_key = normalise_duplicate_text(str(row.get("Programme Name", "")))
        url_key = normalise_url(str(row.get("Source URL", ""))) if row.get("Source URL") else ""
        matched_group: Optional[int] = None
        for group_index, (existing_name, existing_url) in enumerate(signatures):
            same_url = bool(url_key and existing_url and url_key == existing_url)
            similar_name = SequenceMatcher(None, name_key, existing_name).ratio() >= 0.88 if name_key and existing_name else False
            same_parent_page = bool(url_key and existing_url and urlparse(url_key).netloc == urlparse(existing_url).netloc and urlparse(url_key).path.rstrip("/") == urlparse(existing_url).path.rstrip("/"))
            if similar_name and (same_url or same_parent_page):
                matched_group = group_index
                break
        if matched_group is None:
            signatures.append((name_key, url_key))
            groups.append([index])
        else:
            groups[matched_group].append(index)

    duplicate_count = 0
    for group_number, indexes in enumerate(groups, start=1):
        if len(indexes) == 1:
            rows[indexes[0]]["Duplicate Group"] = ""
            rows[indexes[0]]["Duplicate Status"] = "Unique"
            continue
        group_label = f"DUP-{group_number:03d}"
        for position, row_index in enumerate(indexes):
            rows[row_index]["Duplicate Group"] = group_label
            rows[row_index]["Duplicate Status"] = "Primary" if position == 0 else "Duplicate"
            if position > 0:
                duplicate_count += 1
                rows[row_index]["Review Status"] = "Needs human review"
    return duplicate_count


def format_workbook(writer: pd.ExcelWriter, sheet_columns: Dict[str, List[str]]) -> None:
    """Apply lightweight review-friendly formatting to workbook sheets."""
    fills = {
        "Likely valid": PatternFill("solid", fgColor="E2F0D9"),
        "Needs human review": PatternFill("solid", fgColor="FFF2CC"),
        "Low confidence": PatternFill("solid", fgColor="FCE4D6"),
        "Duplicate": PatternFill("solid", fgColor="D9EAF7"),
    }
    for sheet_name, columns in sheet_columns.items():
        worksheet = writer.sheets[sheet_name]
        worksheet.freeze_panes = "A2"
        worksheet.auto_filter.ref = worksheet.dimensions
        for idx, column in enumerate(columns, start=1):
            worksheet.column_dimensions[get_column_letter(idx)].width = min(60, max(12, len(column) + 2))
        if sheet_name == "Extracted Programmes":
            status_index = columns.index("Review Status") if "Review Status" in columns else -1
            duplicate_index = columns.index("Duplicate Status") if "Duplicate Status" in columns else -1
            for row in worksheet.iter_rows(min_row=2):
                status = row[status_index].value if status_index >= 0 else ""
                duplicate = row[duplicate_index].value if duplicate_index >= 0 else ""
                fill = fills.get("Duplicate" if duplicate == "Duplicate" else status)
                if fill:
                    for cell in row:
                        cell.fill = fill



def read_targets(input_path: str) -> List[UniversityTarget]:
    with open(input_path, newline="", encoding="utf-8-sig") as handle:
        reader = csv.DictReader(handle)
        missing = [column for column in INPUT_COLUMNS if column not in (reader.fieldnames or [])]
        if missing:
            raise ValueError(f"{input_path} is missing required column(s): {', '.join(missing)}")
        targets = []
        for row in reader:
            name = str(row.get("University Name", "")).strip()
            seed_url = str(row.get("Seed URL", "")).strip()
            if not name and not seed_url:
                continue
            if not name or not seed_url:
                raise ValueError("Each target row must include University Name and Seed URL.")
            targets.append(UniversityTarget(
                name=name,
                country=str(row.get("Country", "")).strip(),
                seed_url=seed_url,
                allowed_domain=str(row.get("Allowed Domain", "")).strip() or normalise_domain(seed_url),
                notes=str(row.get("Notes", "")).strip(),
            ))
    return targets


def run_scraper(
    input_path: str = INPUT_FILE,
    output_path: str = OUTPUT_FILE,
    progress_callback: Optional[Callable[[str], None]] = None,
    **_: object,
) -> str:
    def report(message: str) -> None:
        print(message)
        if progress_callback:
            progress_callback(message)

    start_time = datetime.now(timezone.utc)
    candidate_rows: List[Dict[str, object]] = []
    programme_rows: List[Dict[str, object]] = []
    warnings: List[str] = []
    error_details = ""
    status = "Completed"
    pages_visited = 0
    universities_processed = 0
    candidate_pages_read = 0
    fallback_rows_written = 0
    universities_with_candidates = 0
    universities_with_extracted_programmes = 0
    duplicate_rows = 0

    try:
        targets = read_targets(input_path)
        report(f"Loaded {len(targets)} university row(s) from {input_path}")
        for target in targets:
            report(f"Crawling {target.name} ({target.allowed_domain})")
            universities_processed += 1
            candidates, visited_count, university_warnings = crawl_university(target)
            pages_visited += visited_count
            if candidates:
                universities_with_candidates += 1
            warnings.extend([f"{target.name}: {warning}" for warning in university_warnings])
            university_programme_rows: List[Dict[str, object]] = []
            robot_parser = get_robot_parser(normalise_url(target.seed_url))
            for page in candidates:
                candidate_rows.append({
                    "University Name": page.university_name,
                    "Country": page.country,
                    "URL": page.url,
                    "Page Title": page.title,
                    "Candidate Score": page.score,
                    "Matched Keywords": ", ".join(page.matched_keywords),
                    "Candidate Type": page.candidate_type,
                    "Reason": page.reason,
                    "Needs Review": page.needs_review,
                })
                try:
                    html = fetch_html(page.url, robot_parser)
                    time.sleep(REQUEST_DELAY_SECONDS)
                    candidate_pages_read += 1
                    if html:
                        programme = extract_programme_from_candidate(target, page, html)
                        if programme:
                            university_programme_rows.append(programme_to_dict(programme))
                except Exception as exc:
                    warnings.append(f"{target.name}: extraction failed for {page.url}: {exc}")
            if candidates and not university_programme_rows:
                fallback_pages = candidates[:min(5, len(candidates))]
                university_programme_rows.extend(programme_to_dict(make_fallback_programme_row(target, page)) for page in fallback_pages)
                fallback_rows_written += len(fallback_pages)
            if university_programme_rows:
                universities_with_extracted_programmes += 1
            programme_rows.extend(university_programme_rows)
            report(f"Finished {target.name}: visited {visited_count}, found {len(candidates)} candidate page(s), wrote {len(university_programme_rows)} programme row(s)")
        if warnings:
            status = "Completed with warnings"
    except Exception as exc:
        status = "Failed"
        error_details = str(exc)

    end_time = datetime.now(timezone.utc)
    duplicate_rows = mark_duplicate_programmes(programme_rows)
    run_log_row = {
        "Start Time": start_time.strftime("%Y-%m-%d %H:%M:%S UTC"),
        "End Time": end_time.strftime("%Y-%m-%d %H:%M:%S UTC"),
        "Universities Processed": universities_processed,
        "Universities With Candidates": universities_with_candidates,
        "Universities With Extracted Programmes": universities_with_extracted_programmes,
        "Total Pages Visited": pages_visited,
        "Total Candidate Pages": len(candidate_rows),
        "Candidate Pages Read": candidate_pages_read,
        "Total Programme Rows": len(programme_rows),
        "Fallback Rows": fallback_rows_written,
        "Duplicate Rows": duplicate_rows,
        "Status": status,
        "Warnings": " | ".join(warnings[:20]),
        "Errors": error_details,
    }

    with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
        pd.DataFrame(candidate_rows, columns=CANDIDATE_COLUMNS).to_excel(writer, sheet_name="Candidate Pages", index=False)
        pd.DataFrame(programme_rows, columns=PROGRAMME_COLUMNS).to_excel(writer, sheet_name="Extracted Programmes", index=False)
        pd.DataFrame([run_log_row], columns=RUN_LOG_COLUMNS).to_excel(writer, sheet_name="Run Log", index=False)
        format_workbook(writer, {
            "Candidate Pages": CANDIDATE_COLUMNS,
            "Extracted Programmes": PROGRAMME_COLUMNS,
            "Run Log": RUN_LOG_COLUMNS,
        })
    report(f"Wrote {output_path}")
    return output_path


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="Collect candidate inbound/short-term programme pages from university websites.")
    parser.add_argument("--input", default=os.getenv("INPUT_FILE", INPUT_FILE), help="Path to target_universities.csv")
    parser.add_argument("--output", default=os.getenv("OUTPUT_FILE", OUTPUT_FILE), help="Path for output_programmes.xlsx")
    return parser.parse_args()


def main() -> None:
    args = parse_args()
    run_scraper(input_path=args.input, output_path=args.output)


if __name__ == "__main__":
    main()
